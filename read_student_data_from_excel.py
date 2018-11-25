#! /usr/bin/env python
# -*- coding: utf8 -*-

# author: dreampython
# date: 2018-10-23
# 将第 0014题中的 student.xls 文件中的内容写到 student.xml 文件中

# 步骤1： 读取students.xlsx中的数据
# 步骤2： 创建xml对象

from openpyxl import load_workbook
from xml.etree.ElementTree import Element, ElementTree
from xml.etree import ElementTree as ET
from xml.dom import minidom
import html

def read_excel_file():
    students = {}
    wb = load_workbook('./students.xlsx')
    sheet = wb['student']
    # 最大行数
    row = sheet.max_row
    # 最大列数
    column = sheet.max_column
    for i in range(1,row+1):
        students[sheet['A'+str(i)].value] = [sheet['B'+str(i)].value,sheet['C'+str(i)].value,\
            sheet['D'+str(i)].value,sheet['E'+str(i)].value]
    
    return students

def create_xml_file(students):
    student_item = {"id": ['名字','数学','语文','英文']}
    str1 = '\n<!--\n' + '\t学生信息表\n' + '\t' + str(student_item) + '\n-->\n'
    str2 = '{\n\t'
    for j,k in students.items():
        str2 += '"' + j + '"' + ' : ' + str(k) + '\n\t'
    str2 = str2[:-1]
    str2 += '}\n'
    return str1 + str2

def write_data_to_xml(students_string):
    root = Element('root')
    students1 = Element('students')
    students1.text = students_string
    root.append(students1)
    rough_string = ET.tostring(root,'utf-8')
    reparsed = minidom.parseString(rough_string)
    with open('./xmltest.xml','w',encoding='utf-8') as f:
        reparsed.writexml(f,indent='',newl='\n',encoding='utf-8')

if __name__ == "__main__":
    student_dict = read_excel_file()
    students_string = create_xml_file(student_dict)
    write_data_to_xml(students_string)
