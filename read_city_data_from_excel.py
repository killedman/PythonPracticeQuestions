#! /usr/bin/env python
# -*- coding: utf8 -*-

# author: dreampython
# date: 2018-10-25
# 第 0018 题： 将 第 0015 题中的 city.xls 文件中的内容写到 city.xml 文件中

from openpyxl import load_workbook
from xml.etree.ElementTree import Element, ElementTree
from xml.etree import ElementTree as ET
from xml.dom import minidom
import html


def read_excel_file():
    citys = {}
    wb = load_workbook('./citys.xlsx')
    sheet = wb['citys']
    # 最大行数
    row = sheet.max_row
    # 最大列数
    column = sheet.max_column
    for i in range(1,row+1):
        citys[sheet['A'+str(i)].value] = [sheet['B'+str(i)].value]
    return citys

def create_xml_file(citys):
    str1 = '\n<!--\n\t城市信息\n-->\n'
    str2 = '{\n\t'
    for j,k in citys.items():
        str2 += '"' + j + '"' + ' : ' + str(k) + '\n\t'
    str2 = str2[:-1]
    str2 += '}\n'
    return str1 + str2

def write_data_to_xml(citys_string):
    root = Element('root')
    citys1 = Element('citys')
    citys1.text = citys_string
    root.append(citys1)
    rough_string = ET.tostring(root,'utf-8')
    reparsed = minidom.parseString(rough_string)
    with open('./xmlcitys.xml','w',encoding='utf-8') as f:
        reparsed.writexml(f,indent='',newl='\n',encoding='utf-8')

if __name__ == "__main__":
    citys_dict = read_excel_file()
    citys_string = create_xml_file(citys_dict)
    write_data_to_xml(citys_string)


