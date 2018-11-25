#! /usr/bin/env python
# -*- coding: utf8 -*-

# author: dreampython
# date: 2018-10-25
# 第0019题：将第 0016题中的numbers.xls文件中的内容写到 numbers.xml

from openpyxl import load_workbook
from xml.etree.ElementTree import Element, ElementTree
from xml.etree import ElementTree as ET
from xml.dom import minidom
import html


def read_excel_file():
    numbers = []
    wb = load_workbook('./numbers.xlsx')
    sheet = wb['numbers']
    # 最大行数
    row = sheet.max_row
    # 最大列数
    column = sheet.max_column
    for i in range(1,row+1):
        numbers.append([sheet['A'+str(i)].value,sheet['B'+str(i)].value,sheet['C'+str(i)].value])
    return numbers

def create_xml_file(numbers):
    str1 = '\n<!--\n\t数字信息\n-->\n'
    str2 = '{\n\t'
    for n in numbers:
        str2 += str(n) + '\n\t'
    str2 = str2[:-1]
    str2 += '}\n'
    return str1 + str2

def write_data_to_xml(numbers_string):
    root = Element('root')
    numbers1 = Element('numbers')
    numbers1.text = numbers_string
    root.append(numbers1)
    rough_string = ET.tostring(root,'utf-8')
    reparsed = minidom.parseString(rough_string)
    with open('./numbers.xml','w',encoding='utf-8') as f:
        reparsed.writexml(f,indent='',newl='\n',encoding='utf-8')

if __name__ == "__main__":
    numbers_list = read_excel_file()
    numbers_string = create_xml_file(numbers_list)
    write_data_to_xml(numbers_string)


